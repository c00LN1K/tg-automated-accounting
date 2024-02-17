import datetime
import json

import openpyxl
import os

import numpy as np
import pandas as pd
from openpyxl.styles import Font

from db import Task

MONTH = ['', 'январь', 'февраль', 'март', 'апрель', 'май', 'июнь', 'июль', 'август', 'сентябрь', 'октябрь', 'ноябрь',
         'декабрь']
WEEKDAY = ['понедельник', 'вторник', 'среда', 'четверг', 'пятница', 'суббота', 'воскресенье']

ROOT = os.getcwd()
active_tasks = {}
last_index = 0
default_data = {'Дата': [], 'Имя': [], 'Время': [], 'Длительность': [], 'Стоимость': [], 'Оплачено': []}


def check_root():
    if os.getcwd() != ROOT:
        os.chdir(ROOT)


def get_filename(date):
    first_day_week = date - datetime.timedelta(days=date.weekday())
    last_day_week = (date + datetime.timedelta(days=6 - date.weekday())).strftime("%d.%m")
    excel_week = f'{first_day_week.strftime("%d.%m")}-{last_day_week}' + '.xlsx'

    return excel_week


def create_excel_file(data, filename):
    df = pd.DataFrame(default_data)
    for el in data:
        df.loc[len(df.index)] = [f'{el.day}.{el.month}.{el.year}', el.name, el.time.strftime('%H:%M'), el.duration,
                                 el.price, el.is_finished]
    df.to_excel(filename, index=False)

    summ = np.sum(df[df['Оплачено'] == True]['Стоимость'])

    wb = openpyxl.load_workbook(filename=filename)
    ws = wb.get_sheet_by_name(wb.sheetnames[0])
    ws['I1'] = 'Сумма'
    ws['I1'].font = Font(bold=True)
    ws['I2'] = summ
    wb.save(filename)
    wb.close()


def create_one_day_excel_file(date: datetime.date):
    filename = f'{date.strftime("%d.%m.%Y")}.xlsx'
    data = Task.get_by_day(date)
    if not data:
        return
    create_excel_file(data, filename)
    return filename


def create_month_excel_file(date: datetime.date):
    filename = f'{MONTH[date.month]}.xlsx'
    data = Task.get_by_month(date)
    if not data:
        return
    create_excel_file(data, filename)
    return filename


def create_year_excel_file(date: datetime.date):
    filename = f'{date.year}.xlsx'
    data = Task.get_by_year(date)
    if not data:
        return
    create_excel_file(data, filename)
    return filename


def create_all_excel_file():
    filename = 'all.xlsx'
    data = Task.get_all_tasks()
    if not data:
        return
    create_excel_file(data, filename)
    return filename
